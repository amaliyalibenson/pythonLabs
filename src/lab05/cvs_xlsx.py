import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    """
    Конвертирует CSV в XLSX.
    Использовать openpyxl ИЛИ xlsxwriter.
    Первая строка CSV — заголовок.
    Лист называется "Sheet1".
    Колонки — автоширина по длине текста (не менее 8 символов).
    FileNotFoundError: Если CSV файл не существует
    ValueError: Если CSV пустой или без заголовка
    """
    # проверка существования файла
    csv_file = Path(csv_path)
    if not csv_file.exists():
        raise FileNotFoundError(f"CSV файл не найден: {csv_path}")

    data = []

    # чтение CSV
    try:
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            for row in reader:
                data.append(row)
    except Exception as e:
        raise ValueError(f"Ошибка чтения CSV: {e}")

    # валидация данных
    if not data:
        raise ValueError("CSV файл пустой")

    if not data[0]:  # Проверка заголовка
        raise ValueError("CSV файл не содержит заголовка")

    # создание Excel книги
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # запись данных
    for row in data:
        ws.append(row)

    # настройка авто-ширины колонок
    for col_num, column_cells in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_num)

        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        # минимальная ширина 10, максимальная 50
        adjusted_width = min(max_length + 2, 50)
        adjusted_width = max(adjusted_width, 10)

        ws.column_dimensions[column_letter].width = adjusted_width

    # сохранение файла
    try:
        wb.save(xlsx_path)
    except Exception as e:
        raise ValueError(f"Ошибка сохранения XLSX: {e}")


csv_to_xlsx("data/samples/people1.csv","data/out/people.xlsx")